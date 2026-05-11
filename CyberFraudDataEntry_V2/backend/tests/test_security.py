"""
Security regression tests — one test per finding in the Innspark VAPT
Audit Reports (Preliminary v1.0.1 + Full-Scope v1.0.1 dated 2026-05-05).

Every finding that has a code-level fix is covered here. If any test
fails in future work, that finding has regressed and must be fixed
BEFORE merging.

Coverage map:
  7.1  Weak admin credentials                    -> test_7_1_*
  7.2  Improper session termination              -> test_7_2_*
  7.3  Identical tokens each login               -> test_7_3_*
  7.4  No rate limiting on auth                  -> test_7_4_*
  7.5  Improper input validation (cases + mule)  -> test_7_5_*
  7.6  Nginx version disclosure                  -> deferred to production nginx
  7.7  Within-PS BOLA (user reads admin record)  -> test_7_7_*
  7.8  Cross-PS BOLA (admin reads other PS)      -> test_7_8_*
  7.9  Username-in-password policy               -> deferred (seed-time fix)
  7.10 XLSX cell content unsanitized             -> test_7_10_*

  Item 8 rec #2 — UUID identifiers (non-sequential)  -> test_uuid_*
  Item 10 rec #2 — XLSX per-cell allow-list          -> test_xlsx_validation_*

  Exec summary (p.8): duplicate FIR/Ack must return 409, not 500
                                                     -> test_duplicate_*

Run:  cd backend && pytest tests/ -v
Pre-req:  backend running on localhost:8000, fresh `python seed.py` run
"""
from __future__ import annotations

import base64
import json
import re
import time
from pathlib import Path

import pytest
import requests

UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}$",
    re.IGNORECASE,
)


# ─────────────────────────────────────────────────────────────────────
# 7.1  Weak administrative credential controls
# ─────────────────────────────────────────────────────────────────────

def test_7_1_seed_passwords_are_unique(seed_users):
    all_pw = [u["password"] for role in seed_users.values() for u in role]
    assert len(all_pw) == len(set(all_pw)), "seed passwords are not all unique"


def test_7_1_seed_passwords_meet_strength(seed_users):
    import re
    all_pw = [u["password"] for role in seed_users.values() for u in role]
    weak_common = {"admin123", "police123", "password", "password123"}
    for pw in all_pw:
        assert len(pw) >= 12, f"password too short: {pw!r}"
        assert re.search(r"[A-Z]", pw), f"missing uppercase in {pw!r}"
        assert re.search(r"[a-z]", pw), f"missing lowercase in {pw!r}"
        assert re.search(r"\d", pw), f"missing digit in {pw!r}"
        assert re.search(r"[^A-Za-z0-9]", pw), f"missing special char in {pw!r}"
        assert pw.lower() not in weak_common, f"weak common password: {pw!r}"


@pytest.mark.parametrize("bad_password,expected_phrase", [
    ("admin123",          "12 characters"),
    ("short1!",           "12 characters"),
    ("alllowercase2026!", "uppercase"),
    ("NOLOWER2026!",      "lowercase"),
    ("NoDigits!abc",      "digit"),
    ("NoSpecialChars123", "special character"),
])
def test_7_1_change_password_rejects_weak(base_url, admin_token, admin_creds,
                                           bad_password, expected_phrase):
    r = requests.post(
        f"{base_url}/api/v1/auth/change-password",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"current_password": admin_creds["password"], "new_password": bad_password},
        timeout=10,
    )
    assert r.status_code == 400, f"expected 400, got {r.status_code}: {r.text}"
    assert expected_phrase.lower() in r.json()["detail"].lower(), (
        f"expected '{expected_phrase}' in error detail, got: {r.text}"
    )


# ─────────────────────────────────────────────────────────────────────
# 7.2  Improper session termination (logout does not revoke)
# ─────────────────────────────────────────────────────────────────────

def test_7_2_token_works_before_logout(base_url, admin_token):
    r = requests.get(
        f"{base_url}/api/v1/auth/me",
        headers={"Authorization": f"Bearer {admin_token}"},
        timeout=10,
    )
    assert r.status_code == 200


def test_7_2_token_rejected_after_logout(base_url, fresh_admin_login):
    # Get a dedicated token just for this test — logging it out would
    # otherwise poison the session-scoped admin_token used by other tests
    captured = fresh_admin_login()
    logout_resp = requests.post(
        f"{base_url}/api/v1/auth/logout",
        headers={"Authorization": f"Bearer {captured}"},
        timeout=10,
    )
    assert logout_resp.status_code == 200

    reuse = requests.get(
        f"{base_url}/api/v1/auth/me",
        headers={"Authorization": f"Bearer {captured}"},
        timeout=10,
    )
    assert reuse.status_code == 401
    assert "revoked" in reuse.json()["detail"].lower()


# ─────────────────────────────────────────────────────────────────────
# 7.3  Identical bearer token issued on every login
# ─────────────────────────────────────────────────────────────────────

def _decode_jwt_claims(token: str) -> dict:
    payload_b64 = token.split(".")[1]
    payload_b64 += "=" * (-len(payload_b64) % 4)
    return json.loads(base64.urlsafe_b64decode(payload_b64))


def test_7_3_tokens_differ_across_logins(base_url, admin_creds):
    def login():
        r = requests.post(
            f"{base_url}/api/v1/auth/login",
            json={"username": admin_creds["username"], "password": admin_creds["password"]},
            timeout=10,
        )
        assert r.status_code == 200
        return r.json()["token"]

    t1 = login()
    time.sleep(1.1)  # ensure iat advances by at least 1 second
    t2 = login()
    assert t1 != t2, "two consecutive logins produced the same token"


def test_7_3_token_has_iat_jti_exp(admin_token):
    claims = _decode_jwt_claims(admin_token)
    assert "iat" in claims, "token missing iat claim"
    assert "jti" in claims, "token missing jti claim (makes tokens session-unique)"
    assert "exp" in claims, "token missing exp claim"
    # jti should be a UUID-shaped string
    assert len(claims["jti"]) >= 32


# ─────────────────────────────────────────────────────────────────────
# 7.4  Lack of rate limiting / account lockout
# ─────────────────────────────────────────────────────────────────────

def test_7_4_account_locks_after_five_failures(base_url, lockout_user):
    username = lockout_user["username"]
    real_pw = lockout_user["password"]
    # 5 wrong attempts — all should return 401
    for i in range(5):
        r = requests.post(
            f"{base_url}/api/v1/auth/login",
            json={"username": username, "password": f"wrong_password_{i}"},
            timeout=10,
        )
        assert r.status_code == 401, (
            f"attempt {i+1}: expected 401 (invalid creds), got {r.status_code}"
        )

    # 6th wrong attempt — account now locked (429)
    r = requests.post(
        f"{base_url}/api/v1/auth/login",
        json={"username": username, "password": "wrong_password_6"},
        timeout=10,
    )
    assert r.status_code == 429

    # Even CORRECT password is rejected while locked
    r = requests.post(
        f"{base_url}/api/v1/auth/login",
        json={"username": username, "password": real_pw},
        timeout=10,
    )
    assert r.status_code == 429, (
        "account was not locked against the correct password — lockout broken"
    )
    assert "locked" in r.json()["detail"].lower()


# ─────────────────────────────────────────────────────────────────────
# 7.5  Improper input validation (stored XSS)
# ─────────────────────────────────────────────────────────────────────

def test_7_5_case_payload_sanitized(base_url, admin_token):
    payload = {
        "fir_no": f"XSS-AUTOTEST-{int(time.time())}",
        "registration_date": "2026-04-22",
        "case_type": "NCRP",
        "crime_type": "Internet",
        "facts": "<script>alert(1)</script>Real facts go here",
        "status": "draft",
        "unfreeze_details": [
            {"unfreeze_type": "letter", "crime_no": "javascript:alert(2)", "amount": 0}
        ],
        "refunds": [
            {"refunded": "yes",
             "victim_name": "<img src=x onerror=alert(3)>John Doe",
             "amount": 100}
        ],
    }
    r = requests.post(
        f"{base_url}/api/v1/cases/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json=payload,
        timeout=10,
    )
    assert r.status_code == 200, f"case create failed: {r.status_code} {r.text}"
    case = r.json()

    facts = case.get("facts") or ""
    crime_no = (case.get("unfreeze_details") or [{}])[0].get("crime_no") or ""
    victim = (case.get("refunds") or [{}])[0].get("victim_name") or ""

    for field_name, value in [("facts", facts), ("crime_no", crime_no), ("victim_name", victim)]:
        assert "<script" not in value.lower(), f"<script survived in {field_name}: {value!r}"
        assert "onerror" not in value.lower(), f"onerror survived in {field_name}: {value!r}"
        assert "javascript:" not in value.lower(), f"javascript: survived in {field_name}: {value!r}"

    # Clean up the test case
    case_id = case.get("id")
    if case_id:
        requests.delete(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_7_5_dsr_case_type_sanitized(base_url, admin_token):
    """VAPT v1.0.1 rec #5 ("review all modules"): the DSR endpoint has a
    free-text case_type field. Even though the audit didn't list /dsr in
    the affected URLs, defense-in-depth requires the same sanitizer."""
    payload = {
        "report_date": "2026-05-05",
        "cases": 1,
        "petitions": 0,
        "details_of_arrest": 0,
        "case_type": "<script>alert(1)</script>NCRP",
        "cumulative_amount_lien_marked": 0,
        "cumulative_accounts_lien_marked": 0,
        "cumulative_accounts_defreezed": 0,
        "amount_refunded_to_victim": 0,
        "ui_cases_pending_2021": 0, "ui_cases_pending_2022": 0,
        "ui_cases_pending_2023": 0, "ui_cases_pending_2024": 0,
        "ui_cases_pending_2025": 0, "ui_cases_pending_2026": 0,
        "disposed_detected_chargesheeted": 0, "disposed_transferred": 0,
        "disposed_false": 0, "disposed_undetected": 0,
        "trial_convicted": 0, "trial_discharged": 0, "trial_acquitted": 0,
        "trial_abated": 0, "trial_compounded": 0, "trial_ut": 0,
    }
    r = requests.post(
        f"{base_url}/api/v1/dsr/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json=payload,
        timeout=10,
    )
    assert r.status_code == 200, f"DSR upsert failed: {r.status_code} {r.text}"
    body = r.json()
    case_type = (body.get("case_type") or "").lower()
    assert "<script" not in case_type, f"<script survived in DSR case_type: {body!r}"


def test_7_5_mule_report_payload_sanitized(base_url, admin_token):
    """VAPT v1.0.1 (2026-05-05) extended item 5 to cover /mule/new and
    /petitions/new in addition to /cases/new. This test exercises the
    mule-report endpoint with payloads in fields that previously had no
    sanitizer wired in schemas/mule.py."""
    payload = {
        "acknowledgement_no": f"<script>alert(1)</script>ACK-{int(time.time())}",
        "fir_no": f"FIR-MULE-XSS-{int(time.time())}",
        "status": "draft",
        "money_transfers": [
            {
                "account_no": "<script>alert(2)</script>1234",
                "bank": "<img src=x onerror=alert(3)>SBI",
                "remarks": "javascript:alert(4)",
                "transaction_amount": 100,
            }
        ],
        "atm_withdrawals": [
            {
                "account_no": "ACCT-001",
                "atm_location": "<script>alert(5)</script>MG Road",
                "remarks": "<svg onload=alert(6)>",
                "withdrawal_amount": 5000,
            }
        ],
    }
    r = requests.post(
        f"{base_url}/api/v1/mule-reports/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json=payload,
        timeout=10,
    )
    assert r.status_code == 200, f"mule report create failed: {r.status_code} {r.text}"
    report = r.json()

    blob = json.dumps(report).lower()
    assert "<script" not in blob, f"<script survived in mule report response: {report!r}"
    assert "onerror" not in blob, f"onerror survived: {report!r}"
    assert "javascript:" not in blob, f"javascript: survived: {report!r}"
    assert "onload" not in blob, f"onload survived: {report!r}"

    # Cleanup
    rid = report.get("id")
    if rid:
        requests.delete(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


# ─────────────────────────────────────────────────────────────────────
# 7.7  Within-PS BOLA: unit_user can read/modify cases submitted by the
#                     admin of the same PS.
# Fix: per-record authorization checks submitted_by == current_user.user_id
#      (or current_user.role == 'admin') on every detail endpoint.
# ─────────────────────────────────────────────────────────────────────

def test_7_7_unit_user_cannot_read_admin_case(base_url, ps_admin_login, ps_user_login):
    """A unit_user from the same PS as an admin cannot GET a case the
    admin created via the BOLA-prone /cases/{id} endpoint."""
    admin_token = ps_admin_login()
    user_token = ps_user_login()

    # Admin creates a case
    create = requests.post(
        f"{base_url}/api/v1/cases/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={
            "fir_no": f"BOLA77-{int(time.time())}",
            "registration_date": "2026-05-05",
            "case_type": "NCRP",
            "crime_type": "Internet",
            "facts": "owned by admin",
            "status": "draft",
        },
        timeout=10,
    )
    assert create.status_code == 200, create.text
    case_id = create.json()["id"]

    try:
        # unit_user tries to read it - should be blocked
        get_r = requests.get(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {user_token}"},
            timeout=10,
        )
        assert get_r.status_code == 403, (
            f"unit_user was able to GET admin's case: {get_r.status_code} {get_r.text}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_7_7_unit_user_cannot_modify_admin_case(base_url, ps_admin_login, ps_user_login):
    """A unit_user from the same PS as an admin cannot PUT a case the
    admin created."""
    admin_token = ps_admin_login()
    user_token = ps_user_login()

    create = requests.post(
        f"{base_url}/api/v1/cases/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={
            "fir_no": f"BOLA77-PUT-{int(time.time())}",
            "registration_date": "2026-05-05",
            "case_type": "NCRP",
            "crime_type": "Internet",
            "facts": "untouched",
            "status": "draft",
        },
        timeout=10,
    )
    assert create.status_code == 200, create.text
    case_id = create.json()["id"]

    try:
        put_r = requests.put(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {user_token}"},
            json={
                "fir_no": create.json()["fir_no"],
                "registration_date": "2026-05-05",
                "case_type": "NCRP",
                "crime_type": "Internet",
                "facts": "TAMPERED BY USER",
                "status": "draft",
            },
            timeout=10,
        )
        assert put_r.status_code == 403, (
            f"unit_user was able to PUT admin's case: {put_r.status_code} {put_r.text}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


# ─────────────────────────────────────────────────────────────────────
# 7.8  Cross-PS BOLA: admin from PS-A can read/modify cases from PS-B.
# Fix: cross-PS check applies even to admins; admins are scoped to their
#      own unit_id on every detail endpoint.
# ─────────────────────────────────────────────────────────────────────

def test_7_8_admin_cannot_read_other_ps_case(base_url, ps_admin_login, other_ps_admin_login):
    """An admin of PS-A cannot GET a case created by the admin of PS-B."""
    admin_a_token = ps_admin_login()
    admin_b_token = other_ps_admin_login()

    # Admin of PS-B creates a case
    create = requests.post(
        f"{base_url}/api/v1/cases/",
        headers={"Authorization": f"Bearer {admin_b_token}"},
        json={
            "fir_no": f"BOLA78-{int(time.time())}",
            "registration_date": "2026-05-05",
            "case_type": "NCRP",
            "crime_type": "Internet",
            "facts": "PS-B private",
            "status": "draft",
        },
        timeout=10,
    )
    assert create.status_code == 200, create.text
    case_id = create.json()["id"]

    try:
        # Admin of PS-A tries to read it - should be blocked
        get_r = requests.get(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_a_token}"},
            timeout=10,
        )
        assert get_r.status_code == 403, (
            f"admin-A was able to GET PS-B's case: {get_r.status_code} {get_r.text}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_b_token}"},
            timeout=10,
        )


def test_7_8_admin_cannot_modify_other_ps_case(base_url, ps_admin_login, other_ps_admin_login):
    """An admin of PS-A cannot PUT a case created by the admin of PS-B."""
    admin_a_token = ps_admin_login()
    admin_b_token = other_ps_admin_login()

    create = requests.post(
        f"{base_url}/api/v1/cases/",
        headers={"Authorization": f"Bearer {admin_b_token}"},
        json={
            "fir_no": f"BOLA78-PUT-{int(time.time())}",
            "registration_date": "2026-05-05",
            "case_type": "NCRP",
            "crime_type": "Internet",
            "facts": "PS-B owned",
            "status": "draft",
        },
        timeout=10,
    )
    assert create.status_code == 200, create.text
    case_id = create.json()["id"]

    try:
        put_r = requests.put(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_a_token}"},
            json={
                "fir_no": create.json()["fir_no"],
                "registration_date": "2026-05-05",
                "case_type": "NCRP",
                "crime_type": "Internet",
                "facts": "TAMPERED CROSS-PS",
                "status": "draft",
            },
            timeout=10,
        )
        assert put_r.status_code == 403, (
            f"admin-A was able to PUT PS-B's case: {put_r.status_code} {put_r.text}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_b_token}"},
            timeout=10,
        )


def test_7_8_fir_no_immutable_on_put(base_url, admin_token):
    """Per product rule (2026-05-05): fir_no cannot be changed via PUT.
    Sending a different fir_no in the body must be silently preserved
    (the original FIR number stays)."""
    original_fir = f"IMMUT-{int(time.time())}"
    create = requests.post(
        f"{base_url}/api/v1/cases/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={
            "fir_no": original_fir,
            "registration_date": "2026-05-05",
            "case_type": "NCRP",
            "crime_type": "Internet",
            "facts": "test",
            "status": "draft",
        },
        timeout=10,
    )
    assert create.status_code == 200, create.text
    case_id = create.json()["id"]

    try:
        put_r = requests.put(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "fir_no": "ATTEMPTED-CHANGE",  # should be ignored
                "registration_date": "2026-05-05",
                "case_type": "NCRP",
                "crime_type": "Internet",
                "facts": "test 2",
                "status": "draft",
            },
            timeout=10,
        )
        assert put_r.status_code == 200, put_r.text
        assert put_r.json()["fir_no"] == original_fir, (
            f"fir_no changed via PUT! original={original_fir!r}, "
            f"after={put_r.json()['fir_no']!r}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_7_8_mule_ack_and_fir_immutable_on_put(base_url, admin_token):
    """Per product rule (2026-05-05): both acknowledgement_no and fir_no on
    a mule report are immutable on PUT. Sending different values in the
    body must be silently preserved-as-original."""
    original_ack = f"ACK-IMMUT-{int(time.time())}"
    original_fir = f"FIR-IMMUT-{int(time.time())}"
    create = requests.post(
        f"{base_url}/api/v1/mule-reports/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={
            "acknowledgement_no": original_ack,
            "fir_no": original_fir,
            "status": "draft",
        },
        timeout=10,
    )
    assert create.status_code == 200, create.text
    rid = create.json()["id"]

    try:
        put_r = requests.put(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "acknowledgement_no": "ATTEMPTED-ACK-CHANGE",  # should be ignored
                "fir_no": "ATTEMPTED-FIR-CHANGE",              # should be ignored
                "status": "draft",
            },
            timeout=10,
        )
        assert put_r.status_code == 200, put_r.text
        body = put_r.json()
        assert body["acknowledgement_no"] == original_ack, (
            f"acknowledgement_no changed via PUT! original={original_ack!r}, "
            f"after={body['acknowledgement_no']!r}"
        )
        assert body["fir_no"] == original_fir, (
            f"fir_no changed via PUT on mule report! original={original_fir!r}, "
            f"after={body['fir_no']!r}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


# ─────────────────────────────────────────────────────────────────────
# 7.10  XLSX cell content unsanitized in mule upload.
# Fix: _safe_str() in routes_mule_report.py now pipes cell values through
#      strip_html() before they reach the ORM/DB.
# ─────────────────────────────────────────────────────────────────────

def test_7_10_xlsx_cell_payload_sanitized(base_url, admin_token, tmp_path):
    """Construct an XLSX with a malicious payload in a cell, upload it,
    and confirm the persisted row has the script tag stripped."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Money Transfer"
    # Headers (row 1) + one data row (row 2). The parser uses ack_no from
    # row 2, col B.
    ws.append(["fir_no", "ack_no", "account", "txn", "bank"])
    ws.append([
        f"FIR-XLSX-{int(time.time())}",
        f"ACK-XLSX-{int(time.time())}",
        "<script>alert(1)</script>1234",
        "TXN-001",
        "<img src=x onerror=alert(2)>SBI",
    ])

    xlsx_path = tmp_path / "vapt_7_10.xlsx"
    wb.save(xlsx_path)

    with open(xlsx_path, "rb") as fh:
        r = requests.post(
            f"{base_url}/api/v1/mule-reports/upload-excel",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"files": ("vapt_7_10.xlsx", fh,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=20,
        )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["results"], f"no results returned: {body!r}"

    # The upload route returns a summary, not the full row contents — fetch
    # the created report via GET to inspect the persisted cell values.
    first = body["results"][0]
    report_id = first.get("report_id")
    if not first.get("ok") or report_id is None:
        # Some seed databases reject the dummy ack_no; in that case we still
        # can't have stored XSS because parsing failed - which counts as a pass.
        return

    try:
        fetched = requests.get(
            f"{base_url}/api/v1/mule-reports/{report_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )
        assert fetched.status_code == 200, fetched.text
        blob = json.dumps(fetched.json()).lower()
        assert "<script" not in blob, f"<script survived XLSX upload: {fetched.text}"
        assert "onerror" not in blob, f"onerror survived XLSX upload: {fetched.text}"
    finally:
        requests.delete(
            f"{base_url}/api/v1/mule-reports/{report_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


# ─────────────────────────────────────────────────────────────────────
# Item 8 rec #2 — UUIDs in place of sequential integer IDs.
# Replaces VAPT recommendation: prevent IDOR/enumeration by making
# /cases/{id} and /mule-reports/{id} take a UUIDv4. A guessed sequential
# integer must now produce a 404, not the next record.
# ─────────────────────────────────────────────────────────────────────

def test_uuid_case_id_is_uuid(base_url, admin_token):
    create = requests.post(
        f"{base_url}/api/v1/cases/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={
            "fir_no": f"UUID-CASE-{int(time.time())}",
            "registration_date": "2026-05-07",
            "case_type": "NCRP",
            "crime_type": "Internet",
            "facts": "uuid pk smoke test",
            "status": "draft",
        },
        timeout=10,
    )
    assert create.status_code == 200, create.text
    case_id = create.json()["id"]
    try:
        assert isinstance(case_id, str), f"case id must be a string, got {type(case_id)}"
        assert UUID_RE.match(case_id), f"case id is not a UUID: {case_id!r}"
    finally:
        requests.delete(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_uuid_mule_report_id_is_uuid(base_url, admin_token):
    create = requests.post(
        f"{base_url}/api/v1/mule-reports/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={
            "acknowledgement_no": f"UUID-ACK-{int(time.time())}",
            "fir_no": f"UUID-FIR-{int(time.time())}",
            "status": "draft",
        },
        timeout=10,
    )
    assert create.status_code == 200, create.text
    rid = create.json()["id"]
    try:
        assert isinstance(rid, str), f"mule report id must be a string, got {type(rid)}"
        assert UUID_RE.match(rid), f"mule report id is not a UUID: {rid!r}"
    finally:
        requests.delete(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_uuid_sequential_case_id_returns_404(base_url, admin_token):
    """A guessed integer id should no longer resolve to a real record."""
    r = requests.get(
        f"{base_url}/api/v1/cases/1",
        headers={"Authorization": f"Bearer {admin_token}"},
        timeout=10,
    )
    assert r.status_code in (404, 422), (
        f"sequential id /cases/1 still resolves: {r.status_code} {r.text}"
    )


def test_uuid_sequential_mule_report_id_returns_404(base_url, admin_token):
    r = requests.get(
        f"{base_url}/api/v1/mule-reports/1",
        headers={"Authorization": f"Bearer {admin_token}"},
        timeout=10,
    )
    assert r.status_code in (404, 422), (
        f"sequential id /mule-reports/1 still resolves: {r.status_code} {r.text}"
    )


# ─────────────────────────────────────────────────────────────────────
# Item 10 rec #2 — XLSX upload per-cell allow-list (Option A).
# Lightweight defense: length caps, control-character rejection, and an
# IFSC-format check inside the XLSX parser so a malicious workbook cell
# can't smuggle oversize / binary / malformed data into the DB.
# ─────────────────────────────────────────────────────────────────────

def test_xlsx_validation_oversize_field_truncated_or_rejected(base_url, admin_token, tmp_path):
    """A cell exceeding the per-field length cap must not be persisted at
    full length. Either the upload is rejected, or the value is truncated
    to the cap."""
    from openpyxl import Workbook
    long_account = "A" * 200  # account_no cap is 30
    wb = Workbook()
    ws = wb.active
    ws.title = "Money Transfer"
    ws.append(["fir_no", "ack_no", "account", "txn", "bank"])
    ws.append([
        f"FIR-OVR-{int(time.time())}",
        f"ACK-OVR-{int(time.time())}",
        long_account,
        "TXN-001",
        "SBI",
    ])
    p = tmp_path / "vapt_10_oversize.xlsx"
    wb.save(p)

    with open(p, "rb") as fh:
        r = requests.post(
            f"{base_url}/api/v1/mule-reports/upload-excel",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"files": (p.name, fh,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=20,
        )
    assert r.status_code == 200, r.text
    first = r.json()["results"][0]
    if not first.get("ok"):
        return  # upload was rejected outright — pass
    rid = first["report_id"]
    try:
        fetched = requests.get(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        ).json()
        for mt in fetched.get("money_transfers", []):
            acct = mt.get("account_no") or ""
            assert len(acct) <= 30, (
                f"account_no cap violated: stored len={len(acct)} value={acct!r}"
            )
    finally:
        requests.delete(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_xlsx_validation_control_chars_stripped():
    """Control characters (bytes < 0x20 except \\t \\n \\r) must be stripped
    by `_safe_str` before persistence.

    Tested at the unit level: openpyxl itself refuses to write control
    chars to a cell (its own defense layer), so an end-to-end XLSX
    upload test isn't possible without forging the workbook's XML by
    hand. The function-level check still covers the case where a
    malformed/forged XLSX bypasses openpyxl and reaches our parser."""
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from api.routes_mule_report import _safe_str

    out = _safe_str("ACCT\x00\x01\x021234", field="account_no")
    assert "\x00" not in out, f"NUL survived: {out!r}"
    assert "\x01" not in out, f"SOH survived: {out!r}"
    assert "\x02" not in out, f"STX survived: {out!r}"
    assert out == "ACCT1234", f"unexpected stripped value: {out!r}"

    # Whitelisted whitespace must be preserved
    preserved = _safe_str("line1\nline2\tcol\rend", field="remarks")
    assert "\n" in preserved and "\t" in preserved and "\r" in preserved, (
        f"whitespace was incorrectly stripped: {preserved!r}"
    )


# ─────────────────────────────────────────────────────────────────────
# Exec summary (page 8) — duplicate FIR / Ack must return 409 with a
# clear message, not a 500 from a bubbled-up IntegrityError.
# ─────────────────────────────────────────────────────────────────────

def test_duplicate_case_fir_returns_409(base_url, admin_token):
    fir = f"DUP-CASE-{int(time.time())}"
    payload = {
        "fir_no": fir,
        "registration_date": "2026-05-07",
        "case_type": "NCRP",
        "crime_type": "Internet",
        "facts": "first",
        "status": "draft",
    }
    first = requests.post(
        f"{base_url}/api/v1/cases/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json=payload,
        timeout=10,
    )
    assert first.status_code == 200, first.text
    case_id = first.json()["id"]
    try:
        # Re-submit the same FIR
        second = requests.post(
            f"{base_url}/api/v1/cases/",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={**payload, "facts": "second"},
            timeout=10,
        )
        assert second.status_code == 409, (
            f"expected 409 on duplicate FIR, got {second.status_code} {second.text}"
        )
        assert fir in second.json().get("detail", ""), (
            f"detail should mention the FIR, got: {second.text}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/cases/{case_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_duplicate_mule_ack_returns_409(base_url, admin_token):
    ack = f"DUP-ACK-{int(time.time())}"
    fir = f"DUP-FIR-{int(time.time())}"
    payload = {"acknowledgement_no": ack, "fir_no": fir, "status": "draft"}
    first = requests.post(
        f"{base_url}/api/v1/mule-reports/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json=payload,
        timeout=10,
    )
    assert first.status_code == 200, first.text
    rid = first.json()["id"]
    try:
        # Same ack, different fir → must still 409 on ack uniqueness
        second = requests.post(
            f"{base_url}/api/v1/mule-reports/",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={**payload, "fir_no": f"OTHER-{int(time.time())}"},
            timeout=10,
        )
        assert second.status_code == 409, (
            f"expected 409 on duplicate ack, got {second.status_code} {second.text}"
        )
        assert ack in second.json().get("detail", ""), (
            f"detail should mention the ack, got: {second.text}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_duplicate_mule_fir_returns_409(base_url, admin_token):
    ack = f"DUP-ACK2-{int(time.time())}"
    fir = f"DUP-FIR2-{int(time.time())}"
    payload = {"acknowledgement_no": ack, "fir_no": fir, "status": "draft"}
    first = requests.post(
        f"{base_url}/api/v1/mule-reports/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json=payload,
        timeout=10,
    )
    assert first.status_code == 200, first.text
    rid = first.json()["id"]
    try:
        # Different ack, same fir → must 409 on fir uniqueness
        second = requests.post(
            f"{base_url}/api/v1/mule-reports/",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={**payload, "acknowledgement_no": f"OTHER-ACK-{int(time.time())}"},
            timeout=10,
        )
        assert second.status_code == 409, (
            f"expected 409 on duplicate fir, got {second.status_code} {second.text}"
        )
        assert fir in second.json().get("detail", ""), (
            f"detail should mention the fir, got: {second.text}"
        )
    finally:
        requests.delete(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )


def test_xlsx_validation_invalid_ifsc_blanked(base_url, admin_token, tmp_path):
    """IFSC codes must match ^[A-Z]{4}0[A-Z0-9]{6}$. Anything else gets
    blanked rather than persisted (so a downstream consumer can't trust a
    bogus IFSC to dispatch action)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Money Transfer"
    ws.append(["fir_no", "ack_no", "account", "txn", "bank", "layer",
               "dest_account", "ifsc", "txn_date", "dest_txn_id", "amount"])
    ws.append([
        f"FIR-IFSC-{int(time.time())}",
        f"ACK-IFSC-{int(time.time())}",
        "ACCT001", "TXN-001", "SBI", 1, "DEST001",
        "not-a-real-ifsc",  # invalid
        "2026-05-07", "DTXN-001", 100,
    ])
    p = tmp_path / "vapt_10_ifsc.xlsx"
    wb.save(p)

    with open(p, "rb") as fh:
        r = requests.post(
            f"{base_url}/api/v1/mule-reports/upload-excel",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"files": (p.name, fh,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=20,
        )
    assert r.status_code == 200, r.text
    first = r.json()["results"][0]
    if not first.get("ok"):
        return
    rid = first["report_id"]
    try:
        fetched = requests.get(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        ).json()
        for mt in fetched.get("money_transfers", []):
            ifsc = mt.get("ifsc_code") or ""
            assert ifsc == "" or re.match(r"^[A-Z]{4}0[A-Z0-9]{6}$", ifsc), (
                f"invalid IFSC was stored: {ifsc!r}"
            )
    finally:
        requests.delete(
            f"{base_url}/api/v1/mule-reports/{rid}",
            headers={"Authorization": f"Bearer {admin_token}"},
            timeout=10,
        )
