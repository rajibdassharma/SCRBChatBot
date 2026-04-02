from __future__ import annotations

import io
import logging
import tempfile
from decimal import Decimal, InvalidOperation
from typing import List

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File

logger = logging.getLogger(__name__)
MAX_EXCEL_SIZE = 10 * 1024 * 1024  # 10 MB
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from database import get_db
from models.mule_report import MuleReport
from models.money_transfer import MoneyTransfer
from models.other_transaction import OtherTransaction
from models.transaction_on_hold import TransactionOnHold
from models.other_less_than_500 import OtherLessThan500
from models.aeps_transaction import AepsTransaction
from models.atm_withdrawal import AtmWithdrawal
from models.unit import Unit
from schemas.mule import MuleReportCreate, MuleReportResponse, MuleReportListItem
from api.deps import get_current_user, require_admin, CurrentUser

router = APIRouter(prefix="/api/v1/mule-reports", tags=["mule-reports"])


# -- Excel parsing helpers -------------------------------------------------

def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _safe_decimal(val) -> Decimal:
    if val is None:
        return Decimal("0")
    try:
        s = str(val).replace(",", "").strip()
        if not s:
            return Decimal("0")
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _safe_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(str(val).strip().rstrip("."))
    except (ValueError, TypeError):
        return 0


def _parse_excel(file_bytes: bytes, filename: str) -> dict:
    """Parse a bank action Excel file into mule report data."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Extract acknowledgement_no from first data row of first sheet
    ack_no = ""
    first_sheet = wb[wb.sheetnames[0]]
    for row in first_sheet.iter_rows(min_row=2, max_row=2, values_only=True):
        ack_no = _safe_str(row[1]) if len(row) > 1 else ""
        break

    result = {
        "acknowledgement_no": ack_no,
        "fir_no": "",
        "money_transfers": [],
        "other_transactions": [],
        "transactions_on_hold": [],
        "others_less_than_500": [],
        "aeps_transactions": [],
        "atm_withdrawals": [],
    }

    # Sheet name mapping (case-insensitive partial match)
    sheet_map = {
        "money transfer": "money_transfers",
        "other": "other_transactions",
        "transaction put on hold": "transactions_on_hold",
        "others less": "others_less_than_500",
        "aeps": "aeps_transactions",
        "withdrawal through atm": "atm_withdrawals",
    }

    logger.debug(f"[ExcelParse] Sheets found: {wb.sheetnames}")
    for sname in wb.sheetnames:
        sname_lower = sname.lower().strip()
        target = None
        for key, val in sheet_map.items():
            if key in sname_lower:
                # "other" matches both "Other" and "Others Less Then 500"
                # so check more specific patterns first
                if key == "other" and ("less" in sname_lower or "aeps" in sname_lower or "atm" in sname_lower or "hold" in sname_lower or "transfer" in sname_lower):
                    continue
                target = val
                break

        if not target:
            logger.debug(f"[ExcelParse] Skipped sheet '{sname}' — no match")
            continue
        logger.debug(f"[ExcelParse] Sheet '{sname}' → {target}")

        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row in rows:
            if not row or not any(c is not None for c in row):
                continue

            if target == "money_transfers":
                result[target].append({
                    "account_no": _safe_str(row[2]) if len(row) > 2 else "",
                    "transaction_id": _safe_str(row[3]) if len(row) > 3 else "",
                    "bank": _safe_str(row[4]) if len(row) > 4 else "",
                    "layer": _safe_int(row[5]) if len(row) > 5 else 1,
                    "dest_account_no": _safe_str(row[6]) if len(row) > 6 else "",
                    "ifsc_code": _safe_str(row[7]) if len(row) > 7 else "",
                    "transaction_date": _safe_str(row[8]) if len(row) > 8 else "",
                    "dest_transaction_id": _safe_str(row[9]) if len(row) > 9 else "",
                    "transaction_amount": _safe_decimal(row[10]) if len(row) > 10 else Decimal("0"),
                    "disputed_amount": _safe_decimal(row[11]) if len(row) > 11 else Decimal("0"),
                    "reference_no": _safe_str(row[12]) if len(row) > 12 else "",
                    "remarks": _safe_str(row[13]) if len(row) > 13 else "",
                    "action_taken_by_bank": _safe_str(row[14]) if len(row) > 14 else "",
                    "date_of_action": _safe_str(row[15]) if len(row) > 15 else "",
                })

            elif target == "other_transactions":
                result[target].append({
                    "account_no": _safe_str(row[2]) if len(row) > 2 else "",
                    "transaction_id": _safe_str(row[3]) if len(row) > 3 else "",
                    "transaction_date": _safe_str(row[4]) if len(row) > 4 else "",
                    "transaction_amount": _safe_decimal(row[5]) if len(row) > 5 else Decimal("0"),
                    "reference_no": _safe_str(row[6]) if len(row) > 6 else "",
                    "remarks": _safe_str(row[7]) if len(row) > 7 else "",
                    "action_taken_by_bank": _safe_str(row[8]) if len(row) > 8 else "",
                    "date_of_action": _safe_str(row[9]) if len(row) > 9 else "",
                })

            elif target == "transactions_on_hold":
                result[target].append({
                    "account_no": _safe_str(row[2]) if len(row) > 2 else "",
                    "transaction_id": _safe_str(row[3]) if len(row) > 3 else "",
                    "hold_date": _safe_str(row[4]) if len(row) > 4 else "",
                    "hold_amount": _safe_decimal(row[5]) if len(row) > 5 else Decimal("0"),
                    "action_taken_by_bank": _safe_str(row[6]) if len(row) > 6 else "",
                    "date_of_action": _safe_str(row[7]) if len(row) > 7 else "",
                    "layer": _safe_int(row[8]) if len(row) > 8 else 0,
                })

            elif target == "others_less_than_500":
                result[target].append({
                    "account_no": _safe_str(row[2]) if len(row) > 2 else "",
                    "transaction_id": _safe_str(row[3]) if len(row) > 3 else "",
                    "reference_no": _safe_str(row[4]) if len(row) > 4 else "",
                    "remarks": _safe_str(row[5]) if len(row) > 5 else "",
                    "action_taken_by_bank": _safe_str(row[6]) if len(row) > 6 else "",
                    "date_of_action": _safe_str(row[7]) if len(row) > 7 else "",
                })

            elif target == "aeps_transactions":
                result[target].append({
                    "account_no": _safe_str(row[2]) if len(row) > 2 else "",
                    "transaction_id": _safe_str(row[3]) if len(row) > 3 else "",
                    "withdrawal_date": _safe_str(row[4]) if len(row) > 4 else "",
                    "withdrawal_amount": _safe_decimal(row[5]) if len(row) > 5 else Decimal("0"),
                    "reference_no": _safe_str(row[6]) if len(row) > 6 else "",
                    "remarks": _safe_str(row[7]) if len(row) > 7 else "",
                    "action_taken_by_bank": _safe_str(row[8]) if len(row) > 8 else "",
                    "date_of_action": _safe_str(row[9]) if len(row) > 9 else "",
                    "layer": _safe_int(row[13]) if len(row) > 13 else 0,
                })

            elif target == "atm_withdrawals":
                result[target].append({
                    "account_no": _safe_str(row[2]) if len(row) > 2 else "",
                    "transaction_id": _safe_str(row[3]) if len(row) > 3 else "",
                    "withdrawal_datetime": _safe_str(row[4]) if len(row) > 4 else "",
                    "withdrawal_amount": _safe_decimal(row[5]) if len(row) > 5 else Decimal("0"),
                    "disputed_amount": _safe_decimal(row[6]) if len(row) > 6 else Decimal("0"),
                    "atm_id": _safe_str(row[7]) if len(row) > 7 else "",
                    "atm_location": _safe_str(row[8]) if len(row) > 8 else "",
                    "reference_no": _safe_str(row[9]) if len(row) > 9 else "",
                    "remarks": _safe_str(row[10]) if len(row) > 10 else "",
                    "action_taken_by_bank": _safe_str(row[11]) if len(row) > 11 else "",
                    "date_of_action": _safe_str(row[12]) if len(row) > 12 else "",
                })

    logger.info(f"[ExcelParse] Parsed: "
          f"transfers={len(result['money_transfers'])}, other={len(result['other_transactions'])}, "
          f"hold={len(result['transactions_on_hold'])}, <500={len(result['others_less_than_500'])}, "
          f"aeps={len(result['aeps_transactions'])}, atm={len(result['atm_withdrawals'])}")
    return result


# -- helpers ---------------------------------------------------------------

def _eager_options():
    return [
        selectinload(MuleReport.money_transfers),
        selectinload(MuleReport.other_transactions),
        selectinload(MuleReport.transactions_on_hold),
        selectinload(MuleReport.others_less_than_500),
        selectinload(MuleReport.aeps_transactions),
        selectinload(MuleReport.atm_withdrawals),
    ]


def _ts(val) -> str | None:
    return str(val) if val else None


def _child_to_dict(obj, exclude=("report", "report_id", "_sa_instance_state")):
    d = {}
    for k, v in obj.__dict__.items():
        if k.startswith("_") or k in exclude:
            continue
        if k == "created_at":
            d[k] = _ts(v)
        elif hasattr(v, "__float__"):
            d[k] = float(v)
        else:
            d[k] = v
    return d


def _validate_submitted_report(body: MuleReportCreate):
    """Raise 422 if a submitted report is missing required fields."""
    if body.status == "submitted":
        errors = []
        if not body.acknowledgement_no:
            errors.append("acknowledgement_no is required when submitting")
        if not body.fir_no:
            errors.append("fir_no is required when submitting")
        if errors:
            raise HTTPException(status_code=422, detail="; ".join(errors))


def _report_to_response(r: MuleReport, unit_name: str | None = None) -> dict:
    return {
        "id": r.id,
        "unit_id": r.unit_id,
        "unit_name": unit_name,
        "acknowledgement_no": r.acknowledgement_no,
        "fir_no": r.fir_no,
        "status": r.status,
        "money_transfers": [_child_to_dict(c) for c in r.money_transfers],
        "other_transactions": [_child_to_dict(c) for c in r.other_transactions],
        "transactions_on_hold": [_child_to_dict(c) for c in r.transactions_on_hold],
        "others_less_than_500": [_child_to_dict(c) for c in r.others_less_than_500],
        "aeps_transactions": [_child_to_dict(c) for c in r.aeps_transactions],
        "atm_withdrawals": [_child_to_dict(c) for c in r.atm_withdrawals],
        "created_at": _ts(r.created_at),
        "updated_at": _ts(r.updated_at),
    }


def _report_to_list_item(r: MuleReport, unit_name: str | None = None) -> dict:
    return {
        "id": r.id,
        "acknowledgement_no": r.acknowledgement_no,
        "fir_no": r.fir_no,
        "status": r.status,
        "unit_name": unit_name,
        "created_at": _ts(r.created_at),
        "money_transfer_count": len(r.money_transfers) if r.money_transfers else 0,
        "other_count": len(r.other_transactions) if r.other_transactions else 0,
        "hold_count": len(r.transactions_on_hold) if r.transactions_on_hold else 0,
        "less_500_count": len(r.others_less_than_500) if r.others_less_than_500 else 0,
        "aeps_count": len(r.aeps_transactions) if r.aeps_transactions else 0,
        "atm_count": len(r.atm_withdrawals) if r.atm_withdrawals else 0,
    }


async def _create_children(report: MuleReport, body: MuleReportCreate, db: AsyncSession):
    for mt in body.money_transfers:
        db.add(MoneyTransfer(report_id=report.id, **mt.model_dump()))
    for ot in body.other_transactions:
        db.add(OtherTransaction(report_id=report.id, **ot.model_dump()))
    for th in body.transactions_on_hold:
        db.add(TransactionOnHold(report_id=report.id, **th.model_dump()))
    for ol in body.others_less_than_500:
        db.add(OtherLessThan500(report_id=report.id, **ol.model_dump()))
    for ae in body.aeps_transactions:
        db.add(AepsTransaction(report_id=report.id, **ae.model_dump()))
    for aw in body.atm_withdrawals:
        db.add(AtmWithdrawal(report_id=report.id, **aw.model_dump()))


async def _get_unit_name(unit_id: int, db: AsyncSession) -> str | None:
    unit = (await db.execute(select(Unit).where(Unit.id == unit_id))).scalar_one_or_none()
    return unit.name if unit else None


# -- POST /api/v1/mule-reports/ ----------------------------------------

@router.post("/", response_model=MuleReportResponse)
async def create_mule_report(
    body: MuleReportCreate,
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    unit_id = current_user.unit_id
    if not unit_id:
        raise HTTPException(status_code=403, detail="No unit assigned to this account.")

    _validate_submitted_report(body)

    report = MuleReport(
        unit_id=unit_id,
        acknowledgement_no=body.acknowledgement_no or None,
        fir_no=body.fir_no or None,
        status=body.status,
        submitted_by=current_user.user_id,
    )
    db.add(report)
    await db.flush()

    await _create_children(report, body, db)
    await db.commit()

    result = (await db.execute(
        select(MuleReport).where(MuleReport.id == report.id).options(*_eager_options())
    )).scalar_one()

    return _report_to_response(result, current_user.unit_name)


# -- GET /api/v1/mule-reports/ ----------------------------------------

@router.get("/", response_model=List[MuleReportListItem])
async def list_mule_reports(
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    unit_id = current_user.unit_id
    if not unit_id:
        raise HTTPException(status_code=403, detail="No unit assigned.")

    reports = (await db.execute(
        select(MuleReport)
        .where(MuleReport.unit_id == unit_id)
        .options(*_eager_options())
        .order_by(MuleReport.created_at.desc())
        .limit(limit)
        .offset(offset)
    )).scalars().all()

    return [_report_to_list_item(r, current_user.unit_name) for r in reports]


# -- GET /api/v1/mule-reports/search (by acknowledgement number) ------
# NOTE: This route MUST be declared BEFORE /{report_id} to avoid path conflict.

@router.get("/search")
async def search_mule_report(
    ack_no: str = Query(...),
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Search for a mule report by acknowledgement number within the user's unit."""
    unit_id = current_user.unit_id
    q = select(MuleReport).where(MuleReport.acknowledgement_no == ack_no).options(*_eager_options())
    if unit_id and current_user.role != "admin":
        q = q.where(MuleReport.unit_id == unit_id)
    report = (await db.execute(q)).scalar_one_or_none()
    if not report:
        return None
    return _report_to_response(report)


# -- GET /api/v1/mule-reports/{id} ------------------------------------

@router.get("/{report_id}", response_model=MuleReportResponse)
async def get_mule_report(
    report_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    report = (await db.execute(
        select(MuleReport).where(MuleReport.id == report_id).options(*_eager_options())
    )).scalar_one_or_none()

    if not report:
        raise HTTPException(status_code=404, detail="Mule report not found")

    if current_user.role != "admin" and report.unit_id != current_user.unit_id:
        raise HTTPException(status_code=403, detail="Access denied")

    unit_name = await _get_unit_name(report.unit_id, db)
    return _report_to_response(report, unit_name)


# -- PUT /api/v1/mule-reports/{id} ------------------------------------

@router.put("/{report_id}", response_model=MuleReportResponse)
async def update_mule_report(
    report_id: int,
    body: MuleReportCreate,
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    report = (await db.execute(
        select(MuleReport).where(MuleReport.id == report_id).options(*_eager_options())
    )).scalar_one_or_none()

    if not report:
        raise HTTPException(status_code=404, detail="Mule report not found")
    if current_user.role != "admin" and report.unit_id != current_user.unit_id:
        raise HTTPException(status_code=403, detail="Access denied")

    _validate_submitted_report(body)

    # Update report fields
    report.acknowledgement_no = body.acknowledgement_no or None
    report.fir_no = body.fir_no or None
    report.status = body.status
    report.submitted_by = current_user.user_id

    # Delete all existing children
    for child in list(report.money_transfers):
        await db.delete(child)
    for child in list(report.other_transactions):
        await db.delete(child)
    for child in list(report.transactions_on_hold):
        await db.delete(child)
    for child in list(report.others_less_than_500):
        await db.delete(child)
    for child in list(report.aeps_transactions):
        await db.delete(child)
    for child in list(report.atm_withdrawals):
        await db.delete(child)

    await db.flush()

    # Recreate children from payload
    await _create_children(report, body, db)
    await db.commit()

    result = (await db.execute(
        select(MuleReport).where(MuleReport.id == report.id).options(*_eager_options())
    )).scalar_one()

    unit_name = await _get_unit_name(result.unit_id, db)
    return _report_to_response(result, unit_name)


# -- DELETE /api/v1/mule-reports/{id} ---------------------------------

@router.delete("/{report_id}")
async def delete_mule_report(
    report_id: int,
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    report = (await db.execute(
        select(MuleReport).where(MuleReport.id == report_id)
    )).scalar_one_or_none()

    if not report:
        raise HTTPException(status_code=404, detail="Mule report not found")
    if current_user.role != "admin" and report.unit_id != current_user.unit_id:
        raise HTTPException(status_code=403, detail="Access denied")

    await db.delete(report)
    await db.commit()
    return {"ok": True, "deleted": report_id}


# -- POST /api/v1/mule-reports/upload-excel --------------------------------

@router.post("/upload-excel")
async def upload_mule_excel(
    files: List[UploadFile] = File(...),
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Upload one or more bank action Excel files. Each creates a mule report."""
    unit_id = current_user.unit_id
    if not unit_id:
        raise HTTPException(status_code=403, detail="No unit assigned to this account.")

    results = []
    for f in files:
        if not f.filename or not f.filename.endswith((".xlsx", ".xls")):
            results.append({"filename": f.filename, "ok": False, "error": "Not an Excel file"})
            continue

        try:
            file_bytes = await f.read()
            if len(file_bytes) > MAX_EXCEL_SIZE:
                results.append({"filename": f.filename, "ok": False, "error": f"File too large. Maximum: {MAX_EXCEL_SIZE // (1024*1024)}MB"})
                continue
            parsed = _parse_excel(file_bytes, f.filename)

            ack_no = parsed["acknowledgement_no"]

            # Check if report with this ack_no already exists
            existing = None
            if ack_no:
                existing = (await db.execute(
                    select(MuleReport).where(MuleReport.acknowledgement_no == ack_no)
                )).scalar_one_or_none()

            if existing:
                results.append({
                    "filename": f.filename,
                    "ok": False,
                    "error": f"Report with Ack No {ack_no} already exists (ID: {existing.id})",
                })
                continue

            # Create the report
            report = MuleReport(
                unit_id=unit_id,
                acknowledgement_no=ack_no or None,
                fir_no=parsed.get("fir_no") or None,
                status="submitted",
                submitted_by=current_user.user_id,
            )
            db.add(report)
            await db.flush()

            # Add children
            for mt in parsed["money_transfers"]:
                db.add(MoneyTransfer(report_id=report.id, **mt))
            for ot in parsed["other_transactions"]:
                db.add(OtherTransaction(report_id=report.id, **ot))
            for th in parsed["transactions_on_hold"]:
                db.add(TransactionOnHold(report_id=report.id, **th))
            for ol in parsed["others_less_than_500"]:
                db.add(OtherLessThan500(report_id=report.id, **ol))
            for ae in parsed["aeps_transactions"]:
                db.add(AepsTransaction(report_id=report.id, **ae))
            for aw in parsed["atm_withdrawals"]:
                db.add(AtmWithdrawal(report_id=report.id, **aw))

            await db.commit()

            total = (
                len(parsed["money_transfers"]) + len(parsed["other_transactions"]) +
                len(parsed["transactions_on_hold"]) + len(parsed["others_less_than_500"]) +
                len(parsed["aeps_transactions"]) + len(parsed["atm_withdrawals"])
            )

            results.append({
                "filename": f.filename,
                "ok": True,
                "report_id": report.id,
                "acknowledgement_no": ack_no,
                "total_transactions": total,
            })

        except Exception as e:
            results.append({"filename": f.filename, "ok": False, "error": str(e)})

    return {"results": results}


# -- POST /api/v1/mule-reports/parse-excel ---------------------------------

@router.post("/parse-excel")
async def parse_mule_excel(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Parse a single Excel file and return the data without saving. For preview before submit."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="Not an Excel file")

    try:
        file_bytes = await file.read()
        parsed = _parse_excel(file_bytes, file.filename)

        # Convert Decimal to float for JSON serialization
        def _convert(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            if isinstance(obj, dict):
                return {k: _convert(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [_convert(i) for i in obj]
            return obj

        return _convert(parsed)
    except Exception as e:
        logger.error(f"Excel parsing error: {e}")
        raise HTTPException(status_code=500, detail="Failed to parse Excel file")
