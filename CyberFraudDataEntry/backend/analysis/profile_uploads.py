#!/usr/bin/env python3
"""Phase 0 — READ-ONLY profiler for uploaded bank statements.

Answers the questions that size and scope the rest of the work, BEFORE
any table is created or any parser is written:

  1. Which bank templates are present, and which few cover most files?
  2. How many transactions does a statement hold?  (sizes the DB)
  3. Is there a TIME column, or only a date?       (decides if F5 is
                                                    minute-level or
                                                    day-level)
  4. Are counterparty account numbers present?     (decides F4's
                                                    match rate)
  5. What share of files extract cleanly?
  6. Which files fail outright?                    (the OCR queue)

SAFETY
------
Reads only. Writes nothing, creates no tables, touches no database.

It prints AGGREGATES ONLY — counts, percentages, and column-header
names. It never prints a transaction, an amount, a name, an account
number, or any other document content. Header names are structural
("Date", "Narration", "Withdrawal Amt.") and are needed to design the
parsers.

    python analysis/profile_uploads.py [--sample 400] [--full]
"""
from __future__ import annotations

import argparse
import os
import random
import re
import sys
import time
import concurrent.futures as cf
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))
BACKEND = os.path.dirname(HERE)
STMT_DIR = os.path.join(BACKEND, "uploads", "statements")

# Issuers seen in Indian cyber-fraud casework. Matched case-insensitively
# against the first page only — enough to identify the template.
BANKS = [
    "State Bank of India", "SBI", "HDFC", "ICICI", "Axis", "Kotak",
    "Punjab National", "PNB", "Bank of Baroda", "Canara", "Union Bank",
    "IndusInd", "Yes Bank", "IDFC", "Bandhan", "Federal Bank",
    "Indian Bank", "Central Bank", "UCO", "Karnataka Bank", "RBL",
    "AU Small Finance", "Airtel Payments", "Paytm Payments", "Fino",
    "India Post", "IPPB", "Bank of India", "Bank of Maharashtra",
    "Karur Vysya", "South Indian Bank", "Jammu", "Saraswat", "Equitas",
]

# A transaction row almost always starts with a date.
DATE_RE = re.compile(
    r"\b(\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}"      # 01-Jan-2026
    r"|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}"            # 01/01/2026
    r"|\d{4}-\d{2}-\d{2})\b"                     # 2026-01-01
)
# HH:MM, optionally with seconds. The presence of this decides F5.
TIME_RE = re.compile(r"\b([01]?\d|2[0-3]):[0-5]\d(:[0-5]\d)?\b")
# 9-18 digit runs — the shape of an Indian account number.
ACCT_RE = re.compile(r"\b\d{9,18}\b")
# UPI/IMPS/NEFT reference, and the UPI handle form.
UTR_RE = re.compile(r"\b(UTR|RRN|REF|TXN)[\s:/-]*[A-Z0-9]{6,}\b", re.I)
UPI_RE = re.compile(r"\b[\w.\-]{2,}@[a-z]{2,}\b", re.I)
ATM_RE = re.compile(r"\b(ATM|ATW|CASH\s*WDL|CASH\s*WITHDRAWAL|NWD)\b", re.I)

#: Cap pages read per PDF. Statements run long; the first 30 pages
#: are ample for profiling and keep the run bounded.
MAX_PAGES = 30

HDR_HINTS = ("date", "narration", "particular", "description", "withdraw",
             "deposit", "debit", "credit", "balance", "chq", "ref", "value")


def detect_bank(text: str) -> str:
    low = text.lower()
    for b in BANKS:
        if b.lower() in low:
            return b
    return "(unidentified)"


def profile_pdf(path: str) -> dict:
    import pdfplumber
    out = {"ok": False, "reason": "", "bank": "(unidentified)", "pages": 0,
           "txn_rows": 0, "has_time": False, "has_acct": False,
           "has_utr": False, "has_upi": False, "has_atm": False,
           "headers": [], "table_rows": 0}
    try:
        with pdfplumber.open(path) as pdf:
            out["pages"] = len(pdf.pages)
            first = pdf.pages[0].extract_text() or ""
            if not first.strip():
                out["reason"] = "no text layer (scanned)"
                return out
            out["bank"] = detect_bank(first)
            # extract_tables() is the expensive call — seconds per page
            # on a long statement. Page 1 is enough to learn the header
            # layout, and plain text answers every other question here.
            for tbl in (pdf.pages[0].extract_tables() or []):
                if not tbl:
                    continue
                head = [str(c or "").strip().lower() for c in tbl[0]]
                if any(any(h in c for h in HDR_HINTS) for c in head):
                    out["headers"] = [c for c in head if c]
                    break
            all_text = []
            for pg in pdf.pages[:MAX_PAGES]:
                all_text.append(pg.extract_text() or "")
            text = "\n".join(all_text)
            out["txn_rows"] = sum(1 for ln in text.splitlines() if DATE_RE.search(ln))
            out["table_rows"] = out["txn_rows"]
            out["has_time"] = bool(TIME_RE.search(text))
            out["has_acct"] = bool(ACCT_RE.search(text))
            out["has_utr"] = bool(UTR_RE.search(text))
            out["has_upi"] = bool(UPI_RE.search(text))
            out["has_atm"] = bool(ATM_RE.search(text))
            out["ok"] = True
    except Exception as exc:                       # noqa: BLE001
        out["reason"] = f"{type(exc).__name__}"
    return out


def profile_excel(path: str) -> dict:
    out = {"ok": False, "reason": "", "bank": "(unidentified)", "pages": 1,
           "txn_rows": 0, "has_time": False, "has_acct": False,
           "has_utr": False, "has_upi": False, "has_atm": False,
           "headers": [], "table_rows": 0}
    try:
        rows: list[list] = []
        if path.lower().endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                rows.append(list(r))
                if i > 3000:
                    break
            wb.close()
        else:
            import xlrd
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            for i in range(min(ws.nrows, 3000)):
                rows.append(ws.row_values(i))
        if not rows:
            out["reason"] = "empty workbook"
            return out
        blob = "\n".join(" ".join(str(c) for c in r if c is not None) for r in rows[:80])
        out["bank"] = detect_bank(blob)
        # Header = first row containing two or more known header words.
        for r in rows[:40]:
            cells = [str(c or "").strip().lower() for c in r]
            hits = sum(1 for c in cells if any(h in c for h in HDR_HINTS))
            if hits >= 2:
                out["headers"] = [c for c in cells if c]
                break
        full = "\n".join(" ".join(str(c) for c in r if c is not None) for r in rows)
        out["txn_rows"] = sum(
            1 for r in rows
            if DATE_RE.search(" ".join(str(c) for c in r if c is not None))
        )
        out["table_rows"] = max(0, len(rows) - 1)
        out["has_time"] = bool(TIME_RE.search(full))
        out["has_acct"] = bool(ACCT_RE.search(full))
        out["has_utr"] = bool(UTR_RE.search(full))
        out["has_upi"] = bool(UPI_RE.search(full))
        out["has_atm"] = bool(ATM_RE.search(full))
        out["ok"] = True
    except Exception as exc:                       # noqa: BLE001
        out["reason"] = f"{type(exc).__name__}"
    return out


def _profile_one(path: str) -> dict:
    """Worker entry point. Must be module-level to be picklable."""
    r = profile_pdf(path) if path.lower().endswith(".pdf") else profile_excel(path)
    r["kind"] = "pdf" if path.lower().endswith(".pdf") else "excel"
    return r


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--sample", type=int, default=400)
    ap.add_argument("--full", action="store_true")
    ap.add_argument("--seed", type=int, default=11)
    ap.add_argument("--workers", type=int, default=12)
    args = ap.parse_args()

    files = sorted(os.listdir(STMT_DIR))
    pdfs = [f for f in files if f.lower().endswith(".pdf")]
    xls = [f for f in files if f.lower().endswith((".xls", ".xlsx"))]
    random.seed(args.seed)
    if not args.full:
        pdfs = random.sample(pdfs, min(args.sample, len(pdfs)))
        xls = random.sample(xls, min(max(args.sample // 2, 1), len(xls)))

    print(f"profiling {len(pdfs):,} PDF + {len(xls):,} Excel "
          f"(corpus: {len([f for f in files if f.lower().endswith('.pdf')]):,} PDF, "
          f"{len([f for f in files if f.lower().endswith((chr(46)+'xls', chr(46)+'xlsx'))]):,} Excel)")
    print("read-only; aggregates only\n")

    # Embarrassingly parallel — one file per worker. This is also a
    # rehearsal for the real pipeline, which parses the same way.
    paths = [os.path.join(STMT_DIR, f) for f in pdfs + xls]
    workers = max(1, min(args.workers, (os.cpu_count() or 4)))
    print(f"  {workers} workers", file=sys.stderr, flush=True)
    results = []
    t0 = time.time()
    with cf.ProcessPoolExecutor(max_workers=workers) as ex:
        for i, r in enumerate(ex.map(_profile_one, paths, chunksize=2), 1):
            results.append(r)
            if i % 25 == 0 or i == len(paths):
                print(f"  {i}/{len(paths)}  ({time.time()-t0:.0f}s)",
                      file=sys.stderr, flush=True)
    elapsed = time.time() - t0

    ok = [r for r in results if r["ok"]]
    bad = [r for r in results if not r["ok"]]

    print("=" * 66)
    print(f"1. EXTRACTION      {len(ok)}/{len(results)} succeeded "
          f"({100*len(ok)/max(1,len(results)):.1f}%)   {elapsed:.0f}s "
          f"({elapsed/max(1,len(results))*1000:.0f} ms/file)")
    for kind in ("pdf", "excel"):
        k = [r for r in results if r["kind"] == kind]
        kok = [r for r in k if r["ok"]]
        if k:
            print(f"     {kind:<6} {len(kok)}/{len(k)} ({100*len(kok)/len(k):.1f}%)")
    if bad:
        print("   failures:")
        for reason, n in Counter(r["reason"] for r in bad).most_common():
            print(f"     {n:>5}  {reason}")

    print(f"\n2. BANK TEMPLATES")
    banks = Counter(r["bank"] for r in ok)
    cum = 0
    for b, n in banks.most_common(15):
        cum += n
        print(f"     {n:>5} ({100*n/len(ok):>5.1f}%)  cum {100*cum/len(ok):>5.1f}%  {b}")
    top3 = sum(n for _, n in banks.most_common(3))
    print(f"   -> top 3 templates cover {100*top3/max(1,len(ok)):.1f}% of parsed files")

    print(f"\n3. TRANSACTIONS PER STATEMENT   (sizes statement_transactions)")
    tx = sorted(r["txn_rows"] for r in ok)
    if tx:
        def pct(p): return tx[min(len(tx) - 1, int(len(tx) * p))]
        mean = sum(tx) / len(tx)
        print(f"     min {tx[0]}   p25 {pct(.25)}   median {pct(.5)}   "
              f"p75 {pct(.75)}   p95 {pct(.95)}   max {tx[-1]}")
        print(f"     mean {mean:.0f}")
        stmts_total = len([f for f in files if f.lower().endswith(('.pdf', '.xls', '.xlsx'))])
        print(f"   -> backfill  ~{mean*stmts_total:,.0f} rows from {stmts_total:,} statements")
        print(f"   -> per year  ~{mean*1450*365:,.0f} rows  (at ~1,450 statements/day)")
        print(f"   -> 3 years   ~{mean*1450*365*3:,.0f} rows")

    print(f"\n4. FIELDS PRESENT   (decides which features are possible)")
    for label, key, feature in [
        ("time of day", "has_time", "F5 velocity -> minute-level"),
        ("account numbers", "has_acct", "F4 counterparty match"),
        ("UTR / ref no", "has_utr", "F4 fallback key"),
        ("UPI handle", "has_upi", "F4 fallback key"),
        ("ATM markers", "has_atm", "F6 cash-out clustering"),
    ]:
        n = sum(1 for r in ok if r[key])
        print(f"     {100*n/max(1,len(ok)):>5.1f}%  {label:<16} {feature}")

    print(f"\n5. COLUMN HEADERS SEEN   (structural, for parser design)")
    hdr = Counter()
    for r in ok:
        for h in r["headers"]:
            if 2 < len(h) < 30:
                hdr[h] += 1
    for h, n in hdr.most_common(20):
        print(f"     {n:>5}  {h}")

    print(f"\n6. PAGES / TABLE ROWS")
    pg = sorted(r["pages"] for r in ok if r["kind"] == "pdf")
    if pg:
        print(f"     pdf pages     median {pg[len(pg)//2]}   max {pg[-1]}")
    tr = sorted(r["table_rows"] for r in ok)
    if tr:
        print(f"     table rows    median {tr[len(tr)//2]}   max {tr[-1]}")
    print("=" * 66)
    return 0


if __name__ == "__main__":
    sys.exit(main())
