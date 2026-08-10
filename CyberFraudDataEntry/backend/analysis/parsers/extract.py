"""Document -> transaction rows.

Three readers, tried in order of how much structure they can rely on:

  read_excel  Excel sheets. Highest yield in the corpus (~90%) because
              the grid is already a grid.
  read_pdf    PDF tables via pdfplumber. Covers 60% of PDFs.
  text lines  Fallback inside read_pdf, for PDFs that have a text layer
              but no table pdfplumber can find — another 10% of PDFs.
              Recovers a median of 47 rows per file.

Scanned PDFs (17%) have no text layer at all and are left for a future
OCR phase. They are recorded as such rather than as failures, because
"needs OCR" is a queue and "failed" is a bug.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, time as dtime

from . import columns as C
from . import values as V

#: Pages read per PDF. The corpus median is 9; the max is 1,787.
#:
#: 200 and not 400: pdfplumber holds parsed objects per page, so memory
#: scales with PAGES, not with file size — the corpus tops out at 4 MB
#: on disk yet a long statement will happily consume gigabytes. That
#: tail of very long files, multiplied across workers, is what
#: exhausted the shared iGPU memory pool and bugchecked the laptop on
#: 2026-08-04 (see analysis/runtime.py).
#:
#: A file that hits the cap is flagged `truncated`, never silently
#: shortened — a partial statement that looks complete would break
#: reconciliation for reasons nobody could find.
MAX_PAGES = 200

#: Beyond this many pages a file is not parsed inline at all; it is
#: reported for a separate, single-worker pass. One 1,787-page document
#: alongside seven other workers is the exact shape of the crash.
DEFER_PAGES = 150

#: Excel rows scanned. Same reasoning.
MAX_XL_ROWS = 60_000


@dataclass
class Txn:
    row_no: int = 0
    txn_date: date | None = None
    txn_time: dtime | None = None
    value_date: date | None = None
    description: str = ""
    ref_no: str | None = None
    debit: float | None = None
    credit: float | None = None
    balance: float | None = None
    #: Balance-chain verdict for THIS row: 1 passed / 0 rejected /
    #: -1 untested. Set by the driver from verify.row_verdicts() after
    #: repair, so it describes the row as stored.
    chain_ok: int = -1
    # Filled by enrich.py from the narration, not read from a column.
    channel: str | None = None
    counterparty_account: str | None = None
    counterparty_name: str | None = None
    counterparty_upi: str | None = None


@dataclass
class Extraction:
    rows: list[Txn] = field(default_factory=list)
    method: str = ""           # table-pdf | text-pdf | excel
    pages: int = 0
    truncated: bool = False
    reason: str = ""           # set only when rows is empty


# --------------------------------------------------------------------
# shared row building
# --------------------------------------------------------------------

def _row_to_txn(cells: list[str], roles: dict[str, int], n: int,
                dayfirst: bool) -> Txn | None:
    """One grid row -> a Txn, or None if it is not a transaction."""
    def cell(role: str) -> str:
        i = roles.get(role, -1)
        return V.clean(cells[i]) if 0 <= i < len(cells) else ""

    d = V.parse_date(cell("txn_date"), dayfirst)
    if d is None:
        return None

    t = Txn(row_no=n, txn_date=d,
            value_date=V.parse_date(cell("value_date"), dayfirst),
            description=cell("description"),
            ref_no=cell("ref_no") or None)

    # Time can live in its own column or share the date cell — HDFC
    # writes "24/09/24 18:05" in one cell.
    t.txn_time = V.parse_time(cell("txn_time")) or V.parse_time(cell("txn_date"))

    if "debit" in roles or "credit" in roles:
        t.debit = V.parse_amount(cell("debit"))
        t.credit = V.parse_amount(cell("credit"))
        # A layout that writes 0.00 in the unused column is not saying
        # "a zero-rupee debit happened"; it is padding. Storing the
        # zero would make every row look like both a debit and a
        # credit to F3's money trail.
        if t.debit == 0.0 and t.credit:
            t.debit = None
        if t.credit == 0.0 and t.debit:
            t.credit = None
    elif "amount" in roles:
        amt = V.parse_amount(cell("amount"))
        flag = V.drcr_flag(cell("drcr")) or V.drcr_flag(cell("amount"))
        if amt is not None:
            if flag == "C":
                t.credit = amt
            elif flag == "D":
                t.debit = amt
            else:
                # Direction unknown at this point. Left on debit as a
                # placeholder and resolved by _infer_direction() from
                # the balance movement, which is the only honest
                # source for it.
                t.debit = amt

    t.balance = V.parse_amount(cell("balance"))
    if t.balance is not None and V.drcr_flag(cell("balance")) == "D":
        # An overdrawn balance shown as "1,200.00Dr".
        t.balance = -t.balance
    return t


def _infer_direction(rows: list[Txn]) -> None:
    """Resolve debit-vs-credit from how the balance moved.

    Applies to single-amount layouts with no Dr/Cr flag, where the
    document simply does not say which way the money went. The balance
    does: if it rose by the amount, it was a credit.

    Deliberately conservative — only rewrites when the movement matches
    the amount to the paisa. A row it cannot explain is left alone and
    fails reconciliation, which is the correct outcome: better a file
    marked unverified than a money trail pointing the wrong way.
    """
    prev = None
    for r in rows:
        if r.balance is not None and prev is not None and r.credit is None:
            amt = r.debit
            if amt:
                delta = r.balance - prev
                if abs(delta - amt) < 0.011:
                    r.debit, r.credit = None, amt
                elif abs(delta + amt) < 0.011:
                    pass                      # already a debit
        if r.balance is not None:
            prev = r.balance


def _merge_continuations(grid: list[list[str]], roles: dict[str, int],
                         dayfirst: bool) -> list[list[str]]:
    """Fold wrapped rows back into the row they belong to.

    A long narration spills onto the next grid row with every other
    cell empty. Left alone those become either junk rows or, worse,
    dropped description text — and the description is where the
    counterparty lives, so losing it costs F4 its match key.
    """
    out: list[list[str]] = []
    di = roles.get("description", -1)
    for cells in grid:
        has_date = V.parse_date(
            cells[roles["txn_date"]] if roles.get("txn_date", -1) < len(cells)
            else "", dayfirst) is not None
        if not has_date and out and 0 <= di < len(cells) and V.clean(cells[di]):
            prev = out[-1]
            while len(prev) <= di:
                prev.append("")
            prev[di] = (prev[di] + " " + V.clean(cells[di])).strip()
            continue
        out.append(list(cells))
    return out


def _build(grid: list[list[str]], roles: dict[str, int]) -> list[Txn]:
    dayfirst = V.infer_dayfirst(
        [r[roles["txn_date"]] for r in grid
         if roles.get("txn_date", 0) < len(r)][:400]
    )
    grid = _merge_continuations(grid, roles, dayfirst)
    rows: list[Txn] = []
    for cells in grid:
        t = _row_to_txn(cells, roles, len(rows), dayfirst)
        if t is not None:
            rows.append(t)
    _infer_direction(rows)
    return rows


# --------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------

def read_excel(path: str) -> Extraction:
    out = Extraction(method="excel", pages=1)
    try:
        if path.lower().endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            raw = []
            for i, r in enumerate(wb[wb.sheetnames[0]].iter_rows(values_only=True)):
                raw.append(["" if c is None else str(c) for c in r])
                if i >= MAX_XL_ROWS:
                    out.truncated = True
                    break
            wb.close()
        else:
            import xlrd
            ws = xlrd.open_workbook(path).sheet_by_index(0)
            n = min(ws.nrows, MAX_XL_ROWS)
            out.truncated = ws.nrows > MAX_XL_ROWS
            raw = [[str(c) for c in ws.row_values(i)] for i in range(n)]
    except Exception as exc:                          # noqa: BLE001
        out.reason = type(exc).__name__
        return out

    hdr = -1
    roles: dict[str, int] = {}
    # Header can sit well down the sheet under an address block; 60
    # rows is past every one seen in this corpus.
    for i, cells in enumerate(raw[:60]):
        if C.is_header_row(cells):
            cand = C.resolve(cells)
            if C.usable(cand):
                hdr, roles = i, cand
                break
    if hdr < 0:
        out.reason = "no usable header row"
        return out
    out.rows = _build(raw[hdr + 1:], roles)
    if not out.rows:
        out.reason = "header found but no transaction rows"
    return out


# --------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------

def _release(page) -> None:
    """Drop pdfplumber's per-page caches.

    This single call is the difference between a bounded run and an
    unbounded one. pdfplumber memoises every page's object list and
    text map on the Page, and the Page is kept alive by the PDF, so
    without an explicit flush a 200-page statement holds all 200 pages
    fully parsed at once — which is how eight workers ate 30 GB.
    """
    try:
        page.flush_cache()
    except Exception:                                  # noqa: BLE001
        pass
    for attr in ("get_textmap", "_get_textmap"):
        fn = getattr(page, attr, None)
        clear = getattr(fn, "cache_clear", None)
        if clear:
            try:
                clear()
            except Exception:                          # noqa: BLE001
                pass


def read_pdf(path: str, defer_pages: int | None = None) -> Extraction:
    # Resolved at CALL time, not from a default argument. A default of
    # DEFER_PAGES binds the value when the function is defined, so the
    # serial pass — which raises the module global to lift the limit
    # for one long document — had no effect at all and every deferred
    # file came back with zero rows.
    if defer_pages is None:
        defer_pages = DEFER_PAGES
    out = Extraction(method="table-pdf")
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            out.pages = len(pdf.pages)
            if out.pages > defer_pages:
                # Not a failure and not an attempt — handed to the
                # single-worker pass so it never runs beside others.
                out.method = "deferred"
                out.reason = f"{out.pages} pages — deferred to serial pass"
                return out
            out.truncated = out.pages > MAX_PAGES

            roles: dict[str, int] = {}
            grid: list[list[str]] = []
            any_text = False
            # Only lines that could BE transactions are kept, not whole
            # pages. On a long statement the discarded material —
            # headers, footers, addresses, terms and conditions — is
            # most of the document.
            cand_lines: list[str] = []

            for idx in range(min(out.pages, MAX_PAGES)):
                pg = pdf.pages[idx]
                try:
                    txt = pg.extract_text() or ""
                    if txt.strip():
                        any_text = True
                        for ln in txt.splitlines():
                            if _LEAD_DATE.match(ln) or (cand_lines and ln.strip()
                                                        and not _MONEY.search(ln)):
                                cand_lines.append(ln)
                    for tbl in (pg.extract_tables() or []):
                        for row in tbl:
                            cells = [V.clean(c) for c in row]
                            if not any(cells):
                                continue
                            if C.is_header_row(cells):
                                cand = C.resolve(cells)
                                # Re-map on every header seen.
                                # Statements repeat the header per
                                # page, and a summary band earlier in
                                # the file must not be allowed to stand
                                # as the mapping.
                                if C.usable(cand):
                                    roles = cand
                                continue
                            if roles:
                                grid.append(cells)
                finally:
                    _release(pg)

            if roles and grid:
                out.rows = _build(grid, roles)

            if not out.rows and not any_text:
                out.method = "scanned"
                out.reason = "no text layer (OCR queue)"
                return out

            # Reconciliation picks the reader, rather than the table
            # reader always winning because it ran first.
            #
            # A PDF table can extract cleanly and still be wrong: when
            # pdfplumber's column boundaries do not match the visual
            # ones, amounts land in neighbouring columns and the rows
            # look perfectly well-formed. Two files in a 40-file sample
            # failed exactly that way — one produced 5 rows from a
            # 5-page statement, the other put debits under credits.
            #
            # The text lines were already collected during this same
            # page pass, so trying the other reader costs no extra I/O
            # and no second parse. Whichever chain actually balances is
            # the one that read the document correctly.
            from .verify import reconcile           # local: avoids a cycle
            best_rows, best = out.rows, reconcile(out.rows) if out.rows else None
            if best is None or not best.verified:
                alt = _read_text_lines(cand_lines)
                if alt:
                    alt_rec = reconcile(alt)
                    better = (
                        best is None
                        or best.rate is None
                        or (alt_rec.rate is not None and alt_rec.rate > best.rate)
                    )
                    if better:
                        best_rows, best = alt, alt_rec
                        out.method = "text-pdf"
            out.rows = best_rows

            if not out.rows:
                out.method = "text-pdf"
                out.reason = "text layer present but no transaction rows"
    except Exception as exc:                          # noqa: BLE001
        out.reason = type(exc).__name__
    return out


#: A transaction line in a table-less PDF: starts with a date and ends
#: with at least two money-shaped tokens. Requiring two is what keeps
#: address lines and summary bands out — they have at most one.
_MONEY = re.compile(r"\d[\d,]*\.\d{2}")
_LEAD_DATE = re.compile(
    r"^\s*(\d{1,2}[-/\s][A-Za-z]{3,4}[-/\s]\d{2,4}"
    r"|\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}"
    r"|\d{4}-\d{2}-\d{2})"
)


def _read_text_lines(lines: list[str]) -> list[Txn]:
    """Fallback for PDFs whose rows are text, not a table.

    Takes candidate LINES, not pages — read_pdf filters as it goes and
    discards the rest of the document rather than holding every page's
    text to the end.

    Column positions are unknowable here, so the shape of the line is
    the only structure available: date at the front, amounts at the
    back, narration in between. Two trailing amounts mean amount +
    balance; three mean debit + credit + balance.

    Direction for the two-amount case is NOT guessed — it is left to
    _infer_direction(), which reads it off the balance movement.
    """
    rows: list[Txn] = []
    dayfirst = V.infer_dayfirst(
        [m.group(1) for ln in lines if (m := _LEAD_DATE.match(ln))][:400]
    )
    for ln in lines:
        m = _LEAD_DATE.match(ln)
        if not m:
            # Continuation of the previous narration.
            if rows and ln.strip() and not _MONEY.search(ln):
                rows[-1].description = (rows[-1].description + " " + ln.strip()).strip()
            continue
        amounts = _MONEY.findall(ln)
        if len(amounts) < 2:
            continue
        d = V.parse_date(m.group(1), dayfirst)
        if d is None:
            continue
        tail = amounts[-3:] if len(amounts) >= 3 else amounts[-2:]
        vals = [float(a.replace(",", "")) for a in tail]
        t = Txn(row_no=len(rows), txn_date=d, txn_time=V.parse_time(ln))
        # Narration = everything between the date and the first of the
        # trailing amounts.
        body = ln[m.end():]
        cut = body.rfind(tail[0])
        t.description = V.clean(body[:cut] if cut > 0 else body)
        if len(vals) == 3:
            t.debit = vals[0] or None
            t.credit = vals[1] or None
            t.balance = vals[2]
        else:
            t.debit = vals[0] or None
            t.balance = vals[1]
        rows.append(t)
    _infer_direction(rows)
    return rows


def read(path: str) -> Extraction:
    if path.lower().endswith((".xls", ".xlsx")):
        return read_excel(path)
    return read_pdf(path)
