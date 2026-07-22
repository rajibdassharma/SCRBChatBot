"""Excel (.xlsx) renderer for the FIR Dashboard → PS-performance table.

Uses openpyxl, which is already a project dependency (mule Excel
parser). Layout matches the on-screen table:

  Row 1  : merged title
  Row 2  : window subtitle
  Row 3  : blank
  Row 4  : bold header row (# | District | Police Station | Total FIRs)
  Row 5+ : data rows
  Last   : Grand Total row

Streams to bytes via an in-memory `BytesIO`; the route wraps that in
a `StreamingResponse` with the openxml content-type.
"""
from __future__ import annotations

import io
from datetime import date as date_t

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schemas.dashboard import FirPsPerformanceRow


# Palette pulls from the same KSP navy/yellow scheme the app uses
# on-screen so the exported file feels like the same product.
_NAVY = "0B2C4A"
_YELLOW = "FFD400"
_LIGHT_BG = "F5F5F7"


def _thin_border() -> Border:
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def render_fir_ps_performance_xlsx(
    rows: list[FirPsPerformanceRow],
    *,
    date_from: date_t,
    date_to: date_t,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "FIR by PS"

    if date_from == date_to:
        window_label = date_from.strftime("%d %b %Y")
    else:
        window_label = (
            f"{date_from.strftime('%d %b %Y')} — {date_to.strftime('%d %b %Y')}"
        )

    # Row 1 — merged title
    ws.cell(row=1, column=1, value="FIR Dashboard — PS Performance")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=_NAVY)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 — window subtitle
    ws.cell(row=2, column=1, value=f"FIRs registered during {window_label}  ·  {len(rows)} police stations")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="666666")

    # Row 3 — blank spacer
    # Row 4 — header
    headers = ["#", "District", "Police Station", "Total FIRs"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 4) else "left", vertical="center")
        cell.border = _thin_border()

    # Row 5+ — data
    for i, r in enumerate(rows, start=1):
        excel_row = 4 + i
        values = [i, r.district, r.ps_name or "—", r.fir_count]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=v)
            cell.border = _thin_border()
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(color="999999")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="right")
                cell.font = Font(bold=True, color=_NAVY if v else "B10000")
            else:
                cell.alignment = Alignment(horizontal="left")
                if col_idx == 2:
                    cell.font = Font(bold=True, color=_NAVY)
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_LIGHT_BG)

    # Grand-total row — bold, yellow highlight
    grand_row = 4 + len(rows) + 1
    ws.cell(row=grand_row, column=1, value="")
    ws.cell(row=grand_row, column=2, value="")
    ws.cell(row=grand_row, column=3, value="Grand Total")
    ws.cell(row=grand_row, column=4, value=sum(r.fir_count for r in rows))
    for col_idx in range(1, 5):
        cell = ws.cell(row=grand_row, column=col_idx)
        cell.font = Font(bold=True, color=_NAVY)
        cell.fill = PatternFill("solid", fgColor=_YELLOW)
        cell.border = _thin_border()
        if col_idx == 3:
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 4:
            cell.alignment = Alignment(horizontal="right")

    # Column widths tuned so a typical district name + PS name fit
    # without wrapping. Widths are in Excel character units.
    widths = [6, 28, 40, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze the header row so long lists stay readable
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
