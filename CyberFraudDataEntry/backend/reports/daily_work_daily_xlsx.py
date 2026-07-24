"""Excel (.xlsx) renderer for the Daily Work Done report (per-PS).

Same shape as daily_work_daily_pdf.py -- one PS per row, three
colour bands (red / yellow / green), 14 metric columns.
"""
from __future__ import annotations

import io
from datetime import date as date_t

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_NAVY       = "0B2C4A"
_NAVY_LIGHT = "1C4267"
_LIGHT_BG   = "F5F5F7"
_YELLOW     = "FFD400"
_RED_BAND   = "B10000"
_YEL_BAND   = "C67C1D"
_GRN_BAND   = "0A6B28"


_GROUPS: list[tuple[str, str, list[tuple[str, str]]]] = [
    ("Notices", _RED_BAND, [
        ("35(3)/41A",            "notices_35_41a_count"),
        ("91/92/94 — Banks",     "notices_91_92_94_banks"),
        ("91/92/94 — Intermed.", "notices_91_92_94_intermediary"),
        ("91/92/94 — Acc. Hldr", "notices_91_92_94_account_holder"),
        ("91/92/94 — CDR/IPDR",  "notices_91_92_94_cdr_ipdr"),
    ]),
    ("Lien / Unlien", _YEL_BAND, [
        ("Lien Req",    "lien_requests_count"),
        ("Freeze Req",  "freeze_requests_count"),
        ("Lien Amount", "total_lien_amount"),
        ("Unlien Req",  "unlien_requests_count"),
        ("Defreeze Req","defreeze_requests_count"),
        ("Unlien Amt",  "total_unlien_amount"),
    ]),
    ("Outcomes", _GRN_BAND, [
        ("Arrests",       "arrests_count"),
        ("Statements",    "statements_count"),
        ("Final (A/B/C)", "final_report_abc"),
    ]),
]

_METRIC_COUNT = sum(len(cols) for _, _, cols in _GROUPS)
_FIXED_COLS = 3
_COL_COUNT = _FIXED_COLS + _METRIC_COUNT


def _thin_border() -> Border:
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def render_daily_work_daily_xlsx(
    rows: list[dict],
    *,
    target_date: date_t,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Work Done"

    # Row 1 -- title
    ws.cell(row=1, column=1, value=f"Daily Work Done — {target_date.strftime('%d %b %Y')}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_COL_COUNT)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=_NAVY)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 -- subtitle
    ws.cell(row=2, column=1,
            value="Per-Police-Station totals across every FIR investigated on this date. "
                  "Amounts in ₹. Blank cells indicate no activity.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_COL_COUNT)
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="666666")

    # Rows 4/5 -- two-row header
    hg_row = 4  # group row
    hm_row = 5  # metric row

    # Fixed columns span both header rows
    for col_idx, label in enumerate(["Sl.No", "Police Station", "FIR Count"], start=1):
        ws.cell(row=hg_row, column=col_idx, value=label)
        ws.merge_cells(start_row=hg_row, start_column=col_idx,
                       end_row=hm_row, end_column=col_idx)
        c = ws.cell(row=hg_row, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()

    col_cursor = _FIXED_COLS + 1
    for gname, band, cols in _GROUPS:
        start_col = col_cursor
        end_col = col_cursor + len(cols) - 1
        ws.cell(row=hg_row, column=start_col, value=gname)
        if end_col > start_col:
            ws.merge_cells(start_row=hg_row, start_column=start_col,
                           end_row=hg_row, end_column=end_col)
        gc = ws.cell(row=hg_row, column=start_col)
        gc.font = Font(bold=True, color="FFFFFF")
        gc.fill = PatternFill("solid", fgColor=band)
        gc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        gc.border = _thin_border()

        for i, (label, _key) in enumerate(cols):
            mc = ws.cell(row=hm_row, column=start_col + i, value=label)
            mc.font = Font(bold=True, color="FFFFFF", size=10)
            mc.fill = PatternFill("solid", fgColor=_NAVY_LIGHT)
            mc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            mc.border = _thin_border()

        col_cursor = end_col + 1

    # Data rows + running totals
    totals: dict[str, float] = {}
    total_fir_count = 0
    total_final: dict[str, int] = {"A": 0, "B": 0, "C": 0}

    for i, r in enumerate(rows, start=1):
        excel_row = hm_row + i
        fir_count = int(r.get("fir_count") or 0)
        total_fir_count += fir_count

        ws.cell(row=excel_row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=excel_row, column=1).border = _thin_border()

        ps_cell = ws.cell(row=excel_row, column=2, value=r.get("ps_name") or "—")
        ps_cell.font = Font(bold=True, color=_NAVY)
        ps_cell.border = _thin_border()
        ps_cell.alignment = Alignment(horizontal="left", vertical="center")

        fc_cell = ws.cell(row=excel_row, column=3,
                          value=fir_count if fir_count else None)
        fc_cell.alignment = Alignment(horizontal="right")
        fc_cell.border = _thin_border()

        col_cursor = _FIXED_COLS + 1
        for _gname, _band, cols in _GROUPS:
            for _label, key in cols:
                if key == "final_report_abc":
                    val = r.get("final_report_abc") or None
                    cell = ws.cell(row=excel_row, column=col_cursor, value=val)
                    for letter in ("A", "B", "C"):
                        total_final[letter] += int(r.get(f"final_report_{letter.lower()}") or 0)
                else:
                    raw = r.get(key)
                    if raw in (None, 0, "0"):
                        cell = ws.cell(row=excel_row, column=col_cursor, value=None)
                    elif "amount" in key:
                        v = float(raw)
                        totals[key] = totals.get(key, 0) + v
                        cell = ws.cell(row=excel_row, column=col_cursor, value=v)
                        cell.number_format = '#,##0'
                    else:
                        v = int(raw)
                        totals[key] = totals.get(key, 0) + v
                        cell = ws.cell(row=excel_row, column=col_cursor, value=v)
                cell.alignment = Alignment(horizontal="right")
                cell.border = _thin_border()
                col_cursor += 1

        if i % 2 == 0:
            for col in range(1, _COL_COUNT + 1):
                target = ws.cell(row=excel_row, column=col)
                if target.fill.fgColor.value in (None, "00000000"):
                    target.fill = PatternFill("solid", fgColor=_LIGHT_BG)

    # Grand-total row
    gt_row = hm_row + len(rows) + 1
    ws.cell(row=gt_row, column=2, value="Grand Total")
    ws.cell(row=gt_row, column=3, value=total_fir_count if total_fir_count else None)
    col_cursor = _FIXED_COLS + 1
    for _gname, _band, cols in _GROUPS:
        for _label, key in cols:
            if key == "final_report_abc":
                parts = [f"{L}:{n}" for L, n in total_final.items() if n]
                ws.cell(row=gt_row, column=col_cursor,
                        value=", ".join(parts) if parts else None)
            elif "amount" in key:
                v = totals.get(key, 0)
                cell = ws.cell(row=gt_row, column=col_cursor, value=v if v else None)
                cell.number_format = '#,##0'
            else:
                v = int(totals.get(key, 0))
                ws.cell(row=gt_row, column=col_cursor, value=v if v else None)
            col_cursor += 1
    for c in range(1, _COL_COUNT + 1):
        cell = ws.cell(row=gt_row, column=c)
        cell.font = Font(bold=True, color=_NAVY)
        cell.fill = PatternFill("solid", fgColor=_YELLOW)
        cell.border = _thin_border()
        if c >= 3:
            cell.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 6   # Sl.No
    ws.column_dimensions["B"].width = 32  # PS name
    ws.column_dimensions["C"].width = 10  # FIR Count
    for c in range(4, _COL_COUNT + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws.freeze_panes = "D6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
