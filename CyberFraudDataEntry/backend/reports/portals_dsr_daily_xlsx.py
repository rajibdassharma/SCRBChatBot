"""Excel (.xlsx) renderer for the Portals DSR daily report.

Same layout + column groups as portals_dsr_daily_pdf.py, but tuned
to openpyxl merge_cells + freeze_panes so the header stays visible
while scrolling long PS lists.
"""
from __future__ import annotations

import io
from datetime import date as date_t

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_NAVY = "0B2C4A"
_NAVY_LIGHT = "1C4267"
_LIGHT_BG = "F5F5F7"
_YELLOW = "FFD400"


# Mirror portals_dsr_daily_pdf.py exactly so the two exports stay in
# lock-step. Any header rewording only needs to happen once, in the
# two lists below.
_GROUPS: list[tuple[str, list[tuple[str, str]]]] = [
    ("NCRP", [
        ("Received",  "ncrp_received"),
        ("Disposed",  "ncrp_disposed"),
        ("Pending",   "ncrp_pending"),
    ]),
    ("Samanvaya", [
        ("Req Recv",   "samanvaya_request_received"),
        ("Actions",    "samanvaya_actions"),
        ("Act Pend",   "samanvaya_action_pending"),
        ("Req Sent",   "samanvaya_request_sent"),
        ("Reply Recv", "samanvaya_reply_received"),
        ("Rep Pend",   "samanvaya_replies_pending"),
    ]),
    ("Sahayog", [
        ("Unlawful",     "sahayog_unlawful_content_removal"),
        ("Intermediary", "sahayog_intermediary_requests"),
        ("Crypto",       "sahayog_crypto_requests"),
    ]),
    ("GRM", [
        ("Req Recv", "grm_request_received"),
        ("Action",   "grm_action"),
        ("Pending",  "grm_pending"),
    ]),
    ("MRM", [
        ("Req Recv", "mrm_request_received"),
        ("Action",   "mrm_action"),
        ("Pending",  "mrm_pending"),
    ]),
    ("Bharatpol", [
        ("Req Sent", "bharatpol_request_received"),
    ]),
    ("OCWC", [
        ("Received", "ocwc_received"),
        ("Disposed", "ocwc_disposed"),
        ("Pending",  "ocwc_pending"),
    ]),
    ("NCMEC (Tipline)", [
        ("Received", "ncmec_received"),
        ("Disposed", "ncmec_disposed"),
        ("Pending",  "ncmec_pending"),
    ]),
]

_METRIC_COUNT = sum(len(cols) for _, cols in _GROUPS)  # 25
_FIXED_COLS = 2  # Sl.No + Police Station
_COL_COUNT = _FIXED_COLS + _METRIC_COUNT  # 27


def _thin_border() -> Border:
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def render_portals_dsr_daily_xlsx(
    rows: list[dict],
    *,
    target_date: date_t,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Portals DSR"

    # Row 1 -- merged title
    ws.cell(row=1, column=1, value=f"Portals DSR — {target_date.strftime('%d %b %Y')}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_COL_COUNT)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=_NAVY)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 -- subtitle
    ws.cell(row=2, column=1,
            value=f"Daily counters across all 8 portals, per Police Station. "
                  f"Submitted rows only; drafts excluded.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_COL_COUNT)
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="666666")

    # Row 4/5 -- two-row header
    header_g_row = 4
    header_m_row = 5

    # Sl.No + Police Station span both header rows
    for col_idx, label in enumerate(["Sl.No", "Police Station"], start=1):
        ws.cell(row=header_g_row, column=col_idx, value=label)
        ws.merge_cells(
            start_row=header_g_row, start_column=col_idx,
            end_row=header_m_row, end_column=col_idx,
        )
        c = ws.cell(row=header_g_row, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()

    # Portal group headers + sub-metric headers
    col_cursor = _FIXED_COLS + 1
    for gname, cols in _GROUPS:
        start_col = col_cursor
        end_col = col_cursor + len(cols) - 1

        ws.cell(row=header_g_row, column=start_col, value=gname)
        if end_col > start_col:
            ws.merge_cells(start_row=header_g_row, start_column=start_col,
                           end_row=header_g_row, end_column=end_col)
        gc = ws.cell(row=header_g_row, column=start_col)
        gc.font = Font(bold=True, color="FFFFFF")
        gc.fill = PatternFill("solid", fgColor=_NAVY)
        gc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        gc.border = _thin_border()

        for i, (label, _key) in enumerate(cols):
            c = ws.cell(row=header_m_row, column=start_col + i, value=label)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=_NAVY_LIGHT)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _thin_border()

        col_cursor = end_col + 1

    # Data rows -- collect running totals per metric key while we go
    totals: dict[str, int] = {}
    for i, r in enumerate(rows, start=1):
        excel_row = header_m_row + i
        ws.cell(row=excel_row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=excel_row, column=1).border = _thin_border()

        ps_cell = ws.cell(row=excel_row, column=2, value=r.get("ps_name") or "—")
        ps_cell.font = Font(bold=True, color=_NAVY)
        ps_cell.border = _thin_border()

        col_cursor = _FIXED_COLS + 1
        for _gname, cols in _GROUPS:
            for label, key in cols:
                v = r.get(key)
                # Blank cell (not 0) when the PS did not submit --
                # matches the paper form convention. Real submitted
                # zeros still render as 0.
                cell = ws.cell(row=excel_row, column=col_cursor,
                               value=None if v is None else int(v))
                cell.alignment = Alignment(horizontal="right")
                cell.border = _thin_border()
                if v is not None:
                    totals[key] = totals.get(key, 0) + int(v)
                col_cursor += 1

        if i % 2 == 0:
            for col in range(1, _COL_COUNT + 1):
                if ws.cell(row=excel_row, column=col).fill.fgColor.value in (None, "00000000"):
                    ws.cell(row=excel_row, column=col).fill = \
                        PatternFill("solid", fgColor=_LIGHT_BG)

    # Grand-total row -- bold yellow band, same treatment as the
    # Daily Work Done report.
    gt_row = header_m_row + len(rows) + 1
    ws.cell(row=gt_row, column=1, value="")
    ws.cell(row=gt_row, column=2, value="Grand Total")
    ws.merge_cells(start_row=gt_row, start_column=1, end_row=gt_row, end_column=2)
    ws.cell(row=gt_row, column=2).alignment = Alignment(horizontal="right", vertical="center")
    col_cursor = _FIXED_COLS + 1
    for _gname, cols in _GROUPS:
        for _label, key in cols:
            n = totals.get(key, 0)
            ws.cell(row=gt_row, column=col_cursor, value=n if n else None)
            col_cursor += 1
    for c in range(1, _COL_COUNT + 1):
        cell = ws.cell(row=gt_row, column=c)
        cell.font = Font(bold=True, color=_NAVY)
        cell.fill = PatternFill("solid", fgColor=_YELLOW)
        cell.border = _thin_border()
        if c >= _FIXED_COLS + 1:
            cell.alignment = Alignment(horizontal="right")

    # Column widths -- narrow metric cols so the whole table stays
    # scannable; wider PS name column since some are ~28 chars.
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    for c in range(3, _COL_COUNT + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9

    # Freeze the two header rows + first two cols so scrolling works.
    ws.freeze_panes = "C6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
