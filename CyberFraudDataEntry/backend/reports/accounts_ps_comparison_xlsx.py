"""Excel (.xlsx) renderer for the Account Details Dashboard →
Per-PS Comparison table.

Mirrors the on-screen table and the PDF sibling — 8 columns:
  # · District · Police Station · Total · <yesterday> · Victim · Mule · Non-Mule
"""
from __future__ import annotations

import io
from datetime import date as date_t, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schemas.dashboard import AccountsPsComparison


_NAVY = "0B2C4A"
_YELLOW = "FFD400"
_LIGHT_BG = "F5F5F7"


def _thin_border() -> Border:
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def render_accounts_ps_comparison_xlsx(
    rows: list[AccountsPsComparison],
    *,
    target_date: date_t,
) -> bytes:
    yesterday = target_date - timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts by PS"

    # Row 1 — title
    ws.cell(row=1, column=1, value="Account Details — PS Comparison")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=_NAVY)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 — subtitle
    subtitle = (
        f"Cumulative as of {target_date.strftime('%d %b %Y')}  ·  "
        f"Yesterday column = {yesterday.strftime('%d %b %Y')}  ·  "
        f"{len(rows)} police stations"
    )
    ws.cell(row=2, column=1, value=subtitle)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="666666")

    # Row 4 — header
    headers = [
        "#", "District", "Police Station", "Total",
        yesterday.strftime("%d %b %Y"),
        "Victim", "Mule", "Non-Mule",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.alignment = Alignment(
            horizontal="center" if col_idx == 1 else ("right" if col_idx >= 4 else "left"),
            vertical="center",
        )
        cell.border = _thin_border()

    # Row 5+ — data
    for i, r in enumerate(rows, start=1):
        excel_row = 4 + i
        values = [
            i, r.unit_name, r.ps_name or "—",
            r.total, r.yesterday_count,
            r.victims, r.mules, r.non_mules,
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=v)
            cell.border = _thin_border()
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(color="999999")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="left")
                cell.font = Font(bold=True, color=_NAVY)
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="left")
                cell.font = Font(color=_NAVY)
            else:
                cell.alignment = Alignment(horizontal="right")
                if col_idx == 4:  # Total
                    cell.font = Font(bold=True, color=_NAVY if v else "B10000")
                elif col_idx == 5:  # Yesterday
                    cell.font = Font(bold=True, color="0A6B28" if v else "999999")
                elif col_idx == 6:  # Victim
                    cell.font = Font(color="0A6B28")
                elif col_idx == 7:  # Mule
                    cell.font = Font(color="8B1919")
                elif col_idx == 8:  # Non-Mule
                    cell.font = Font(color="5B6B7A")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_LIGHT_BG)

    # Grand-total row
    grand_row = 4 + len(rows) + 1
    ws.cell(row=grand_row, column=3, value="Grand Total")
    ws.cell(row=grand_row, column=4, value=sum(r.total for r in rows))
    ws.cell(row=grand_row, column=5, value=sum(r.yesterday_count for r in rows))
    ws.cell(row=grand_row, column=6, value=sum(r.victims for r in rows))
    ws.cell(row=grand_row, column=7, value=sum(r.mules for r in rows))
    ws.cell(row=grand_row, column=8, value=sum(r.non_mules for r in rows))
    for col_idx in range(1, 9):
        cell = ws.cell(row=grand_row, column=col_idx)
        cell.font = Font(bold=True, color=_NAVY)
        cell.fill = PatternFill("solid", fgColor=_YELLOW)
        cell.border = _thin_border()
        if col_idx == 3:
            cell.alignment = Alignment(horizontal="right")
        elif col_idx >= 4:
            cell.alignment = Alignment(horizontal="right")

    # Column widths
    widths = [6, 24, 32, 12, 14, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
